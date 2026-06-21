import json, os, uuid
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile
from sqlalchemy.orm import Session
from pydantic import BaseModel

from orm_utils import orm_to_dict
from database import get_db
from models import Subject, Chapter, Question, QuestionVersion, User
from auth import get_current_user, require_role
from routers.audit_router import log_operation

router = APIRouter(prefix="/api/questions", tags=["题库管理"])

UPLOAD_DIR = "uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)


class QuestionOut(BaseModel):
    id: int; subject_id: int; subject_name: str = ""; chapter_id: int | None = None; chapter_name: str = ""
    specialty: str = ""; type: str; difficulty: int; content: str; options: str; answer: str
    explanation: str = ""; reference: str = ""; images: str = "[]"; is_active: bool; version: int = 1
    created_at: str | None = None; updated_at: str | None = None
    model_config = {"from_attributes": True}

class QuestionCreate(BaseModel):
    subject_id: int; chapter_id: int | None = None; specialty: str = ""
    type: str = "single"; difficulty: int = 1
    content: str; options: str = "[]"; answer: str
    explanation: str = ""; reference: str = ""; images: str = "[]"


@router.get("", response_model=dict)
def list_questions(
    subject_id: int = Query(None), chapter_id: int = Query(None),
    specialty: str = Query(None), type: str = Query(None),
    difficulty: int = Query(None), keyword: str = Query(None),
    page: int = Query(1, ge=1), page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db), _: User = Depends(get_current_user),
):
    q = db.query(Question).filter(Question.is_active == True)
    if subject_id: q = q.filter(Question.subject_id == subject_id)
    if chapter_id: q = q.filter(Question.chapter_id == chapter_id)
    if specialty: q = q.filter(Question.specialty == specialty)
    if type: q = q.filter(Question.type == type)
    if difficulty: q = q.filter(Question.difficulty == difficulty)
    if keyword: q = q.filter(Question.content.contains(keyword))

    total = q.count()
    items = q.order_by(Question.id.desc()).offset((page - 1) * page_size).limit(page_size).all()
    result = [_q_out(item) for item in items]
    return {"total": total, "page": page, "page_size": page_size, "items": result}


@router.get("/export-template")
def export_template(_=Depends(get_current_user)):
    """Download Excel import template"""
    from fastapi.responses import StreamingResponse
    import openpyxl
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "题库导入模板"
    ws.append(["科目", "章节", "题型", "难度(1-5)", "题干", "选项(格式:A.内容;B.内容;C.内容)", "答案", "解析", "规范引用"])
    ws.append(['示例: 建筑检测', '第一章', '单选', 2, '普通混凝土立方体抗压强度标准值用什么表示?', 'A.fcu,k;B.fc;C.fck;D.fcu', 'A', 'fcu,k为立方体抗压强度标准值', 'GB/T 50081'])
    ws.append(['示例: 建筑检测', '第二章', '多选', 3, '以下哪些属于建筑材料检测项目?', 'A.抗压强度检测;B.抗折强度检测;C.坍落度检测;D.含泥量检测', 'ABCD', '建筑材料检测涵盖力学性能和物理性能', 'GB/T 50081'])
    ws.append(['示例: 建筑检测', '第一章', '判断', 1, '水泥越细早期强度越高', '', 'A', '水泥细度影响水化速度和早期强度', 'GB/T 1346'])
    ws.append(['示例: 建筑检测', '第三章', '综合', 4, '简述混凝土配合比设计步骤', '', '一、确定配制强度 二、确定水灰比 三、确定用水量 四、确定水泥用量 五、确定砂率 六、确定砂石用量', '配合比设计需考虑强度和耐久性', 'GB/T 50080'])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=question_import_template.xlsx"})

@router.get('/export-excel')
def export_excel(db: Session = Depends(get_db), _: User = Depends(require_role('admin', 'teacher'))):
    from fastapi.responses import StreamingResponse
    import openpyxl
    from io import BytesIO
    questions = db.query(Question).filter(Question.is_active == True).order_by(Question.id).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '题库导出'
    ws.append(['科目', '章节', '题型', '难度', '题干', '选项', '答案', '解析', '规范引用'])
    type_label = {'single': '单选', 'multi': '多选', 'truefalse': '判断', 'composite': '综合', 'blank': '填空'}
    for q in questions:
        subject_name = q.subject.name if q.subject else ''
        chapter_name = q.chapter.name if q.chapter else ''
        opts = q.options or '[]'
        try:
            opts_list = json.loads(opts)
            opts_str = ';'.join(chr(123)+"o.get('label','')+'.'+o.get('text','')"+chr(125) for o in opts_list) if opts_list else ''
        except:
            opts_str = opts
        ws.append([subject_name, chapter_name, type_label.get(q.type, q.type), q.difficulty, q.content, opts_str, q.answer, q.explanation, q.reference])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': 'attachment; filename=questions_export.xlsx'})


@router.get("/{question_id}", response_model=QuestionOut)
def get_question(question_id: int, db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    q = db.query(Question).get(question_id)
    if not q: raise HTTPException(404, "题目不存在")
    return _q_out(q)


@router.post("", response_model=QuestionOut)
def create_question(data: QuestionCreate, db: Session = Depends(get_db), _: User = Depends(require_role("admin", "teacher"))):
    if not db.query(Subject).get(data.subject_id): raise HTTPException(404, "科目不存在")
    q = Question(**data.model_dump())
    db.add(q); db.commit(); db.refresh(q)
    return _q_out(q)


@router.put("/{question_id}", response_model=QuestionOut)
def update_question(question_id: int, data: QuestionCreate, db: Session = Depends(get_db), _: User = Depends(require_role("admin", "teacher"))):
    q = db.query(Question).get(question_id)
    if not q: raise HTTPException(404, "题目不存在")
    # Save version history
    ver = QuestionVersion(
        question_id=q.id, version=q.version,
        content=q.content, options=q.options, answer=q.answer,
        explanation=q.explanation, changed_by=_.fullname or _.username,
    )
    db.add(ver)
    for k, v in data.model_dump().items():
        setattr(q, k, v)
    q.version += 1
    q.updated_at = datetime.now(timezone.utc)
    db.commit(); db.refresh(q)
    return _q_out(q)


@router.delete("/{question_id}")
def delete_question(question_id: int, db: Session = Depends(get_db), _: User = Depends(require_role("admin", "teacher"))):
    q = db.query(Question).get(question_id)
    if not q: raise HTTPException(404, "题目不存在")
    q.is_active = False
    db.commit()
    return {"ok": True}


@router.get("/{question_id}/versions")
def get_versions(question_id: int, db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    versions = db.query(QuestionVersion).filter(QuestionVersion.question_id == question_id).order_by(QuestionVersion.version.desc()).all()
    return [{"id": v.id, "version": v.version, "content": v.content, "options": v.options, "answer": v.answer, "explanation": v.explanation, "changed_by": v.changed_by, "created_at": v.created_at.isoformat() if v.created_at else None} for v in versions]


# --- Image Upload ---

@router.post("/upload-image")
async def upload_question_image(
    file: UploadFile,
    _: User = Depends(require_role("admin", "teacher")),
):
    """Upload question image, returns URL"""
    ALLOWED = {"png", "jpg", "jpeg", "gif", "webp", "bmp"}
    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in ALLOWED:
        raise HTTPException(400, f"Unsupported image format: .{ext}, supported: {', '.join(sorted(ALLOWED))}")
    contents = await file.read()
    if len(contents) > 10 * 1024 * 1024:
        raise HTTPException(400, "Image must be <= 10MB")
    safe_name = f"{uuid.uuid4().hex}.{ext}"
    filepath = os.path.join(UPLOAD_DIR, safe_name)
    with open(filepath, "wb") as f: f.write(contents)
    return {"url": f"/uploads/{safe_name}"}


# --- Batch Import ---

class BatchImportItem(BaseModel):
    subject_id: int; chapter_id: int | None = None; specialty: str = ""
    type: str = "single"; difficulty: int = 1
    content: str; options: str = "[]"; answer: str
    explanation: str = ""; reference: str = ""

class BatchImportRequest(BaseModel):
    questions: list[BatchImportItem]


@router.post("/batch-import")
def batch_import(data: BatchImportRequest, db: Session = Depends(get_db), _: User = Depends(require_role("admin", "teacher"))):
    count = 0
    for item in data.questions:
        q = Question(**item.model_dump())
        db.add(q); count += 1
    db.commit()
    return {"ok": True, "count": count}


@router.post("/batch-import-excel")
def batch_import_excel(file: UploadFile, db: Session = Depends(get_db), _: User = Depends(require_role("admin", "teacher"))):
    """Import questions from Excel"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file.file)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        count = 0
        subject_cache = {}
        chapter_cache = {}

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]: continue
            subject_name = str(row[0]).strip() if row[0] else "默认科目"
            if subject_name not in subject_cache:
                subj = db.query(Subject).filter(Subject.name == subject_name).first()
                if not subj:
                    subj = Subject(name=subject_name)
                    db.add(subj); db.flush()
                subject_cache[subject_name] = subj.id

            subject_id = subject_cache[subject_name]
            chapter_id = None
            chapter_name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            if chapter_name:
                cache_key = f"{subject_id}:{chapter_name}"
                if cache_key not in chapter_cache:
                    ch = db.query(Chapter).filter(Chapter.subject_id == subject_id, Chapter.name == chapter_name).first()
                    if not ch:
                        ch = Chapter(subject_id=subject_id, name=chapter_name)
                        db.add(ch); db.flush()
                    chapter_cache[cache_key] = ch.id
                chapter_id = chapter_cache[cache_key]

            qtype_map = {"单选": "single", "多选": "multi", "判断": "truefalse", "综合": "composite"}
            qtype = qtype_map.get(str(row[2]).strip() if len(row) > 2 and row[2] else "", "single")
            difficulty = int(row[3]) if len(row) > 3 and row[3] else 1
            content = str(row[4]) if len(row) > 4 and row[4] else ""
            options_raw = str(row[5]).strip() if len(row) > 5 and row[5] else ""
            if options_raw.startswith("["):
                options_text = options_raw
            else:
                parsed_opts = []
                for part in options_raw.replace(";", "\n").split("\n"):
                    part = part.strip()
                    if not part:
                        continue
                    if "." in part[:4]:
                        label, text = part.split(".", 1)
                        parsed_opts.append({"label": label.strip(), "text": text.strip()})
                    else:
                        parsed_opts.append({"label": chr(65 + len(parsed_opts)), "text": part})
                options_text = json.dumps(parsed_opts, ensure_ascii=False)
            answer = str(row[6]) if len(row) > 6 and row[6] else ""
            explanation = str(row[7]) if len(row) > 7 and row[7] else ""
            reference = str(row[8]) if len(row) > 8 and row[8] else ""

            q = Question(subject_id=subject_id, chapter_id=chapter_id, type=qtype, difficulty=difficulty, content=content, options=options_text, answer=answer, explanation=explanation, reference=reference)
            db.add(q); count += 1

        db.commit()
        return {"ok": True, "count": count}
    except Exception as e:
        raise HTTPException(400, f"Import failed: {str(e)}")




def _q_out(q: Question) -> QuestionOut:
    d = orm_to_dict(q)
    # Handle DB NULL values that should be defaults
    for k, default in [("specialty", ""), ("reference", ""), ("images", "[]"), ("version", 1)]:
        if d.get(k) is None:
            d[k] = default
    out = QuestionOut.model_validate(d)
    out.subject_name = q.subject.name if q.subject else ""
    out.chapter_name = q.chapter.name if q.chapter else ""
    return out
